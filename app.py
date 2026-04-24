"""
Sara's Work Management System — Flask Backend
Serves the frontend and connects it to PostgreSQL
"""

from flask import Flask, request, jsonify, render_template, abort, session, send_from_directory
from flask_socketio import SocketIO
from functools import wraps
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from openpyxl import load_workbook
import psycopg2  # type: ignore[import-not-found]
import os, re, json, datetime, secrets, sys, shutil, configparser
from psycopg2.extras import RealDictCursor  # type: ignore[import-not-found]

app = Flask(__name__)
_IS_CLOUD = os.getenv("RENDER", "").strip().lower() == "true"

# ── Persistent secret key (survives restarts; file created on first run) ──────
_SECRET_KEY_FILE = os.path.join(
    os.path.dirname(sys.executable) if getattr(sys, 'frozen', False)
    else os.path.dirname(os.path.abspath(__file__)),
    "secret.key"
)
_env_secret = os.getenv("SECRET_KEY", "").strip()
if _env_secret:
    app.secret_key = _env_secret
else:
    try:
        with open(_SECRET_KEY_FILE, "r", encoding="utf-8") as _f:
            _saved_key = _f.read().strip()
        if _saved_key:
            app.secret_key = _saved_key
        else:
            raise ValueError("empty key file")
    except Exception:
        app.secret_key = secrets.token_hex(32)
        with open(_SECRET_KEY_FILE, "w", encoding="utf-8") as _f:
            _f.write(str(app.secret_key or ""))

app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['SESSION_COOKIE_SECURE'] = _IS_CLOUD or os.getenv("SESSION_COOKIE_SECURE", "").strip().lower() == "true"

# ── CORS: restrict to the server's own origin ────────────────────────────────
_cfg = configparser.ConfigParser()
_cfg.read(os.path.join(
    os.path.dirname(sys.executable) if getattr(sys, 'frozen', False)
    else os.path.dirname(os.path.abspath(__file__)),
    "config.ini"
))
_server_url = _cfg.get("server", "url", fallback="http://127.0.0.1:5000").strip().rstrip("/")
_cors_env = os.getenv("CORS_ORIGINS", "").strip()
if _cors_env:
    _cors_origins = [origin.strip().rstrip("/") for origin in _cors_env.split(",") if origin.strip()]
else:
    _cors_origins = list({_server_url, "http://127.0.0.1:5000", "http://localhost:5000"})

# ── Choose async mode: prefer eventlet if installed ──────────────────────────
try:
    import eventlet as _ev  # noqa: F401
    _async_mode = "eventlet"
except ImportError:
    _async_mode = "threading"

socketio = SocketIO(app, cors_allowed_origins=_cors_origins, async_mode=_async_mode)

# ── Rate limiter (login brute-force protection) ───────────────────────────────
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
limiter = Limiter(get_remote_address, app=app, default_limits=[])

# When bundled as .exe, persist data next to the executable, not inside
# the temporary _MEIPASS extraction folder which is wiped on exit.
if getattr(sys, 'frozen', False):
    _DATA_DIR = os.path.dirname(sys.executable)
else:
    _DATA_DIR = os.path.dirname(os.path.abspath(__file__))

def _pg_config_value(key, env_key, fallback):
    return os.getenv(env_key, _cfg.get("database", key, fallback=fallback)).strip()


PG_CONFIG = {
    "host": _pg_config_value("host", "PGHOST", "127.0.0.1"),
    "port": int(_pg_config_value("port", "PGPORT", "5432") or "5432"),
    "dbname": _pg_config_value("name", "PGDATABASE", "sara_wms"),
    "user": _pg_config_value("user", "PGUSER", "postgres"),
    "password": _pg_config_value("password", "PGPASSWORD", ""),
    "sslmode": _pg_config_value("sslmode", "PGSSLMODE", "prefer"),
}

DATABASE_URL = os.getenv("DATABASE_URL", "").strip()

UPLOAD_FOLDER = os.path.join(_DATA_DIR, "uploads")
MAX_UPLOAD_BYTES = 10 * 1024 * 1024


_ALLOWED_TABLES = frozenset({"app_users", "work_orders", "customers"})


class _QueryResult:
    def __init__(self, rows):
        self._rows = rows
        self._index = 0

    def fetchone(self):
        if self._index >= len(self._rows):
            return None
        row = self._rows[self._index]
        self._index += 1
        return row

    def fetchall(self):
        if self._index >= len(self._rows):
            return []
        rows = self._rows[self._index:]
        self._index = len(self._rows)
        return rows


def _adapt_query(query, params):
    adapted = query
    adapted = adapted.replace("datetime('now')", "CURRENT_TIMESTAMP::text")
    adapted = adapted.replace("date('now')", "CURRENT_DATE::text")
    adapted = adapted.replace("COLLATE NOCASE", "")
    if isinstance(params, dict):
        adapted = re.sub(r"(?<!:):([A-Za-z_][A-Za-z0-9_]*)", r"%(\1)s", adapted)
    else:
        adapted = adapted.replace("?", "%s")
    return adapted


class _PgConnectionAdapter:
    def __init__(self, conn):
        self._conn = conn

    def execute(self, query, params=None):
        adapted = _adapt_query(query, params)
        with self._conn.cursor(cursor_factory=RealDictCursor) as cursor:
            cursor.execute(adapted, params)
            rows = cursor.fetchall() if cursor.description else []
        return _QueryResult(rows)

    def executemany(self, query, seq_params):
        rows = list(seq_params)
        if not rows:
            return
        adapted = _adapt_query(query, rows[0])
        with self._conn.cursor() as cursor:
            cursor.executemany(adapted, rows)

    def commit(self):
        self._conn.commit()

    def rollback(self):
        self._conn.rollback()

    def close(self):
        self._conn.close()

def _table_count(conn, table_name):
    if table_name not in _ALLOWED_TABLES:
        return 0
    try:
        row = conn.execute(f"SELECT COUNT(*) AS total FROM {table_name}").fetchone()
        return int(row["total"]) if row else 0
    except psycopg2.Error:
        return 0


def _hydrate_frozen_runtime_data():
    if not getattr(sys, 'frozen', False):
        return

    bundled_uploads = os.path.join(getattr(sys, '_MEIPASS', _DATA_DIR), "uploads")

    os.makedirs(UPLOAD_FOLDER, exist_ok=True)
    if os.path.isdir(bundled_uploads):
        for file_name in os.listdir(bundled_uploads):
            src = os.path.join(bundled_uploads, file_name)
            dst = os.path.join(UPLOAD_FOLDER, file_name)
            if os.path.isfile(src) and not os.path.exists(dst):
                shutil.copy2(src, dst)

# ── DB helper ─────────────────────────────────────────────────────────────────
def get_db():
    connect_kwargs = dict(PG_CONFIG)
    if DATABASE_URL:
        conn = psycopg2.connect(DATABASE_URL)
    else:
        if not connect_kwargs.get("password"):
            connect_kwargs.pop("password", None)
        conn = psycopg2.connect(**connect_kwargs)
    return _PgConnectionAdapter(conn)


def _ensure_upload_folder():
    os.makedirs(UPLOAD_FOLDER, exist_ok=True)


def _ensure_attachment_schema():
    conn = get_db()
    try:
        row = conn.execute(
            """
            SELECT 1
            FROM information_schema.columns
            WHERE table_schema = 'public' AND table_name = %s AND column_name = %s
            """,
            ("work_orders", "attachment_path"),
        ).fetchone()
        if not row:
            conn.execute("ALTER TABLE work_orders ADD COLUMN attachment_path TEXT")
            conn.commit()
    finally:
        conn.close()


def _normalize_customer_name(value):
    return " ".join(str(value or "").strip().split())


def _normalize_customer_code(value):
    return " ".join(str(value or "").strip().split())


def _ensure_customers_table():
    conn = get_db()
    try:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS customers (
                customer_id BIGSERIAL PRIMARY KEY,
                code        TEXT,
                name        TEXT NOT NULL,
                created_at  TEXT NOT NULL DEFAULT (CURRENT_TIMESTAMP::text),
                updated_at  TEXT NOT NULL DEFAULT (CURRENT_TIMESTAMP::text)
            )
        """)
        row = conn.execute(
            """
            SELECT 1
            FROM information_schema.columns
            WHERE table_schema = 'public' AND table_name = %s AND column_name = %s
            """,
            ("customers", "code"),
        ).fetchone()
        if not row:
            conn.execute("ALTER TABLE customers ADD COLUMN code TEXT")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_customers_name ON customers(name)")
        conn.execute(
            "CREATE UNIQUE INDEX IF NOT EXISTS idx_customers_name_unique_ci "
            "ON customers ((LOWER(name)))"
        )
        conn.execute(
            "CREATE UNIQUE INDEX IF NOT EXISTS idx_customers_code_unique "
            "ON customers ((LOWER(code))) WHERE code IS NOT NULL AND TRIM(code) <> ''"
        )
        conn.commit()
    finally:
        conn.close()


def _sync_customers_from_orders():
    conn = get_db()
    try:
        rows = conn.execute(
            "SELECT DISTINCT customer FROM work_orders WHERE TRIM(COALESCE(customer, '')) <> ''"
        ).fetchall()
        names = [
            (_normalize_customer_name(row["customer"]),)
            for row in rows
            if _normalize_customer_name(row["customer"])
        ]
        if names:
            conn.executemany(
                "INSERT INTO customers (name, created_at, updated_at) "
                "VALUES (?, CURRENT_TIMESTAMP::text, CURRENT_TIMESTAMP::text) ON CONFLICT DO NOTHING",
                names,
            )
            conn.commit()
    finally:
        conn.close()


def _upsert_customer(conn, name, code=None):
    customer_name = _normalize_customer_name(name)
    customer_code = _normalize_customer_code(code)
    if not customer_name:
        return False, False

    row_by_code = None
    if customer_code:
        row_by_code = conn.execute(
            "SELECT customer_id, code, name FROM customers WHERE LOWER(code) = LOWER(?)",
            (customer_code,),
        ).fetchone()
        if row_by_code:
            name_changed = _normalize_customer_name(row_by_code["name"]) != customer_name
            code_changed = _normalize_customer_code(row_by_code["code"]) != customer_code
            if name_changed or code_changed:
                conn.execute(
                    "UPDATE customers SET code = ?, name = ?, updated_at = CURRENT_TIMESTAMP::text WHERE customer_id = ?",
                    (customer_code, customer_name, row_by_code["customer_id"]),
                )
                return False, True
            return False, False

    row_by_name = conn.execute(
        "SELECT customer_id, code, name FROM customers WHERE LOWER(name) = LOWER(?)",
        (customer_name,),
    ).fetchone()
    if row_by_name:
        existing_code = _normalize_customer_code(row_by_name["code"])
        if customer_code and existing_code != customer_code:
            conn.execute(
                "UPDATE customers SET code = ?, updated_at = CURRENT_TIMESTAMP::text WHERE customer_id = ?",
                (customer_code, row_by_name["customer_id"]),
            )
            return False, True
        return False, False

    conn.execute(
        "INSERT INTO customers (code, name, created_at, updated_at) "
        "VALUES (?, ?, CURRENT_TIMESTAMP::text, CURRENT_TIMESTAMP::text)",
        (customer_code or None, customer_name),
    )
    return True, False


def _find_customer_column(headers, aliases):
    for index, header in enumerate(headers):
        normalized = _normalize_customer_name(header).lower()
        if normalized in aliases:
            return index
    return None


def _current_user_is_admin():
    user_key = session.get('user_key')
    if not user_key:
        return False
    conn = get_db()
    try:
        row = conn.execute(
            "SELECT is_admin FROM app_users WHERE user_key = ?", (user_key,)
        ).fetchone()
        return bool(row and row['is_admin'])
    finally:
        conn.close()


def _attachment_display_name(stored_name):
    if not stored_name:
        return None
    base_name = os.path.basename(stored_name)
    if "___" in base_name:
        return base_name.split("___", 1)[1]
    return base_name


def _remove_attachment_file(stored_name):
    if not stored_name:
        return
    file_path = os.path.join(UPLOAD_FOLDER, os.path.basename(stored_name))
    if os.path.exists(file_path):
        os.remove(file_path)


def _ensure_schema():
    """Run schema.sql idempotently so all tables exist on first launch."""
    if getattr(sys, 'frozen', False):
        schema_path = os.path.join(getattr(sys, '_MEIPASS', _DATA_DIR), 'schema.sql')
    else:
        schema_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'schema.sql')
    if not os.path.exists(schema_path):
        return
    with open(schema_path, 'r', encoding='utf-8') as f:
        schema_sql = f.read()
    conn = get_db()
    try:
        customer_table_exists = conn.execute(
            """
            SELECT 1
            FROM information_schema.tables
            WHERE table_schema = 'public' AND table_name = %s
            """,
            ("customers",),
        ).fetchone()
        if customer_table_exists:
            code_column_exists = conn.execute(
                """
                SELECT 1
                FROM information_schema.columns
                WHERE table_schema = 'public' AND table_name = %s AND column_name = %s
                """,
                ("customers", "code"),
            ).fetchone()
            if not code_column_exists:
                conn.execute("ALTER TABLE customers ADD COLUMN code TEXT")
        for statement in [part.strip() for part in schema_sql.split(';') if part.strip()]:
            conn.execute(statement)
        conn.commit()
    finally:
        conn.close()


# ── Bootstrap app_users table (runs once on startup) ─────────────────────────
def _bootstrap():
    conn = get_db()
    try:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS app_users (
                user_key      TEXT PRIMARY KEY,
                name          TEXT NOT NULL,
                role          TEXT NOT NULL,
                password_hash TEXT NOT NULL,
                is_admin      INTEGER NOT NULL DEFAULT 0,
                sections_json TEXT NOT NULL DEFAULT '[]',
                created_at    TEXT NOT NULL DEFAULT (CURRENT_TIMESTAMP::text)
            )
        """)
        count_row = conn.execute("SELECT COUNT(*) AS total FROM app_users").fetchone()
        total_users = int(count_row["total"]) if count_row else 0
        if total_users == 0:
            conn.execute(
                "INSERT INTO app_users VALUES (?,?,?,?,?,?,CURRENT_TIMESTAMP::text)",
                ('admin1', 'Admin 1', 'Administrator',
                 generate_password_hash('Admin@1234'), 1, '[]')
            )
            conn.execute(
                "INSERT INTO app_users VALUES (?,?,?,?,?,?,CURRENT_TIMESTAMP::text)",
                ('admin2', 'Admin 2', 'Administrator',
                 generate_password_hash('Admin@5678'), 1, '[]')
            )
        conn.commit()
    finally:
        conn.close()

_hydrate_frozen_runtime_data()
_ensure_schema()
_bootstrap()
_ensure_attachment_schema()
_ensure_customers_table()
_sync_customers_from_orders()
_ensure_upload_folder()

# ── Auth helpers ──────────────────────────────────────────────────────────────
def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_key' not in session:
            return jsonify({'error': 'Unauthorized'}), 401
        return f(*args, **kwargs)
    return decorated

def _user_to_dict(row):
    return {
        'key':      row['user_key'],
        'name':     row['name'],
        'role':     row['role'],
        'isAdmin':  bool(row['is_admin']),
        'sections': json.loads(row['sections_json'] or '[]'),
    }


@app.route("/health", methods=["GET"])
def health_check():
    """Simple health endpoint for Render checks and uptime monitoring."""
    conn = get_db()
    try:
        conn.execute("SELECT 1")
        return jsonify({"ok": True, "database": "up"})
    except Exception as exc:
        return jsonify({"ok": False, "database": "down", "error": str(exc)}), 503
    finally:
        conn.close()

# ── Serve frontend ────────────────────────────────────────────────────────────
@app.route("/")
def index():
    return render_template("index.html")

# ── Login / session ───────────────────────────────────────────────────────────
@app.route("/api/login", methods=["POST"])
@limiter.limit("10 per minute")
def api_login():
    data = request.get_json(force=True) or {}
    username = data.get('username', '').strip()
    password = data.get('password', '')
    if not username or not password:
        return jsonify({'error': 'Missing credentials'}), 400
    conn = get_db()
    try:
        row = conn.execute(
            "SELECT * FROM app_users WHERE user_key = ?", (username,)
        ).fetchone()
    finally:
        conn.close()
    if not row or not check_password_hash(row['password_hash'], password):
        return jsonify({'error': 'Invalid username or password'}), 401
    session['user_key'] = username
    return jsonify(_user_to_dict(row))


@app.route("/api/logout", methods=["POST"])
def api_logout():
    session.clear()
    return jsonify({'ok': True})


@app.route("/api/session", methods=["GET"])
def api_session():
    user_key = session.get('user_key')
    if not user_key:
        return jsonify({'error': 'Not authenticated'}), 401
    conn = get_db()
    try:
        row = conn.execute(
            "SELECT * FROM app_users WHERE user_key = ?", (user_key,)
        ).fetchone()
    finally:
        conn.close()
    if not row:
        session.clear()
        return jsonify({'error': 'User not found'}), 401
    return jsonify(_user_to_dict(row))

# ══════════════════════════════════════════════════════════════════════════════
#  APP USERS API  (login/permissions users — stored in app_users table)
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/api/login-users", methods=["GET"])
def get_login_users():
    """Public — returns only the minimum info needed to populate the login dropdown."""
    conn = get_db()
    try:
        rows = conn.execute(
            "SELECT user_key, name, is_admin FROM app_users ORDER BY is_admin DESC, name"
        ).fetchall()
        return jsonify([
            {"key": r["user_key"], "name": r["name"], "isAdmin": bool(r["is_admin"])}
            for r in rows
        ])
    finally:
        conn.close()


@app.route("/api/app-users", methods=["GET"])
@login_required
def get_app_users():
    """Returns full user records including sections (requires auth)."""
    conn = get_db()
    try:
        rows = conn.execute(
            "SELECT * FROM app_users ORDER BY is_admin DESC, name"
        ).fetchall()
        return jsonify([_user_to_dict(r) for r in rows])
    finally:
        conn.close()


@app.route("/api/app-users", methods=["POST"])
@login_required
def add_app_user():
    data     = request.get_json(force=True) or {}
    key      = (data.get('key') or '').strip()
    name     = (data.get('name') or '').strip()
    role     = (data.get('role') or 'Operator').strip()
    password = data.get('password', '')
    is_admin = 1 if (data.get('isAdmin') or role == 'Administrator') else 0
    sections = data.get('sections', [])
    if not key or not name or not password:
        abort(400, "key, name and password are required")
    conn = get_db()
    try:
        if conn.execute(
            "SELECT user_key FROM app_users WHERE user_key = ?", (key,)
        ).fetchone():
            return jsonify({'error': 'Username already exists'}), 409
        conn.execute(
            "INSERT INTO app_users VALUES (?,?,?,?,?,?,datetime('now'))",
            (key, name, role, generate_password_hash(password),
             is_admin, json.dumps(sections))
        )
        conn.commit()
        socketio.emit("users_changed", {})
        return jsonify({'ok': True, 'key': key}), 201
    finally:
        conn.close()


@app.route("/api/app-users/<user_key>", methods=["PUT"])
@login_required
def update_app_user(user_key):
    data = request.get_json(force=True) or {}
    conn = get_db()
    try:
        row = conn.execute(
            "SELECT * FROM app_users WHERE user_key = ?", (user_key,)
        ).fetchone()
        if not row:
            return jsonify({'error': 'User not found'}), 404
        name     = (data.get('name') or row['name']).strip()
        role     = (data.get('role') or row['role']).strip()
        is_admin = 1 if (data.get('isAdmin', bool(row['is_admin'])) or role == 'Administrator') else 0
        sections = data.get('sections', json.loads(row['sections_json'] or '[]'))
        password = data.get('password', '').strip()
        if password:
            conn.execute(
                "UPDATE app_users SET name=?,role=?,is_admin=?,sections_json=?,"
                "password_hash=? WHERE user_key=?",
                (name, role, is_admin, json.dumps(sections),
                 generate_password_hash(password), user_key)
            )
        else:
            conn.execute(
                "UPDATE app_users SET name=?,role=?,is_admin=?,sections_json=?"
                " WHERE user_key=?",
                (name, role, is_admin, json.dumps(sections), user_key)
            )
        conn.commit()
        socketio.emit("users_changed", {})
        return jsonify({'ok': True})
    finally:
        conn.close()


@app.route("/api/app-users/<user_key>", methods=["DELETE"])
@login_required
def delete_app_user(user_key):
    conn = get_db()
    try:
        row = conn.execute(
            "SELECT is_admin FROM app_users WHERE user_key = ?", (user_key,)
        ).fetchone()
        if not row:
            return jsonify({'error': 'User not found'}), 404
        if row['is_admin']:
            return jsonify({'error': 'Cannot delete admin users'}), 403
        conn.execute("DELETE FROM app_users WHERE user_key = ?", (user_key,))
        conn.commit()
        socketio.emit("users_changed", {})
        return jsonify({'ok': True})
    finally:
        conn.close()


@app.route("/api/customers", methods=["GET"])
@login_required
def get_customers():
    conn = get_db()
    try:
        rows = conn.execute(
            "SELECT customer_id, code, name, updated_at FROM customers ORDER BY LOWER(name)"
        ).fetchall()
        return jsonify([
            {
                'id': row['customer_id'],
                'code': row['code'] or '',
                'name': row['name'],
                'updatedAt': row['updated_at'],
            }
            for row in rows
        ])
    finally:
        conn.close()


@app.route("/api/customers/import", methods=["POST"])
@login_required
def import_customers():
    if not _current_user_is_admin():
        return jsonify({'error': 'Only administrators can import customers'}), 403

    upload = request.files.get('file')
    if not upload or not upload.filename:
        return jsonify({'error': 'No Excel file was provided'}), 400

    filename = secure_filename(upload.filename)
    lower_name = filename.lower()
    if not lower_name.endswith(('.xlsx', '.xlsm', '.xltx', '.xltm')):
        return jsonify({'error': 'Only Excel .xlsx files are supported'}), 400

    try:
        upload.stream.seek(0)
        workbook = load_workbook(upload.stream, read_only=True, data_only=True)
        sheet = workbook.worksheets[0]
    except Exception:
        return jsonify({'error': 'Could not read the Excel file'}), 400

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return jsonify({'error': 'The Excel sheet is empty'}), 400

    name_header_keywords = {'customer', 'customer name', 'customername', 'customer master', 'name'}
    code_header_keywords = {'code', 'customer code', 'customercode', 'cust code', 'customer id', 'customerid'}
    selected_col = None
    code_col = None
    start_row = 0
    first_row = rows[0]
    selected_col = _find_customer_column(first_row, name_header_keywords)
    code_col = _find_customer_column(first_row, code_header_keywords)
    if selected_col is not None or code_col is not None:
        start_row = 1

    if selected_col is None:
        for candidate in range(max(len(first_row), 1)):
            sample = []
            for row in rows[:25]:
                if candidate < len(row):
                    value = _normalize_customer_name(row[candidate])
                    if value:
                        sample.append(value)
            if sample:
                selected_col = candidate
                break

    if selected_col is None:
        return jsonify({'error': 'No customer name column was found'}), 400

    parsed_customers = []
    seen = set()
    skipped = 0
    for row in rows[start_row:]:
        if selected_col >= len(row):
            skipped += 1
            continue
        name = _normalize_customer_name(row[selected_col])
        code = _normalize_customer_code(row[code_col]) if code_col is not None and code_col < len(row) else ''
        if not name:
            skipped += 1
            continue
        if len(name) > 500:
            name = name[:500].rstrip()
        if len(code) > 100:
            code = code[:100].rstrip()
        key = f"{code.casefold()}::{name.casefold()}" if code else name.casefold()
        if key in seen:
            skipped += 1
            continue
        seen.add(key)
        parsed_customers.append({'code': code, 'name': name})

    if not parsed_customers:
        return jsonify({'error': 'No customer names were found in the Excel file'}), 400

    conn = get_db()
    try:
        inserted = 0
        updated = 0
        for customer in parsed_customers:
            created, changed = _upsert_customer(conn, customer['name'], customer['code'])
            inserted += 1 if created else 0
            updated += 1 if changed else 0
        conn.commit()
        socketio.emit("customers_changed", {})
        return jsonify({
            'ok': True,
            'imported': inserted + updated,
            'inserted': inserted,
            'updated': updated,
            'total': len(parsed_customers),
            'skipped': skipped,
            'columnIndex': selected_col,
            'codeColumnIndex': code_col,
        })
    finally:
        conn.close()

# ══════════════════════════════════════════════════════════════════════════════
#  WORK ORDERS API
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/api/orders", methods=["GET"])
@login_required
def get_orders():
    conn = get_db()
    try:
        rows = conn.execute(
            "SELECT data_json, attachment_path FROM work_orders ORDER BY saved_at DESC"
        ).fetchall()
        orders = []
        for row in rows:
            record = json.loads(row["data_json"])
            record["attachmentPath"] = row["attachment_path"] or None
            record["attachmentName"] = _attachment_display_name(row["attachment_path"])
            orders.append(record)
        return jsonify(orders)
    finally:
        conn.close()


@app.route("/api/orders", methods=["POST"])
@login_required
def save_order():
    record = request.get_json(force=True)
    if not record or not record.get("id"):
        abort(400, "Missing record id")
    if "/" in record["id"]:
        return jsonify({"error": 'W/O # cannot contain "/" — use "-" instead (e.g. SMTPP8-26-27)'}), 400
    now = datetime.datetime.utcnow().isoformat(timespec="seconds")
    conn = get_db()
    try:
        customer_name = _normalize_customer_name(record.get("customer", ""))
        conn.execute("""
            INSERT INTO work_orders
                (id, wo, customer, division, priority, coating, electrode,
                 order_qty, order_value, po_advance, trail, design,
                 store_blanks, shortage, bin_number, rcvd_status, rcvd_date,
                 rcvd_qty, del_start, del_weeks, delivery_date, po_entry_date,
                 data_json, saved_by, saved_at, updated_at)
            VALUES
                (:id,:wo,:customer,:division,:priority,:coating,:electrode,
                 :order_qty,:order_value,:po_advance,:trail,:design,
                 :store_blanks,:shortage,:bin_number,:rcvd_status,:rcvd_date,
                 :rcvd_qty,:del_start,:del_weeks,:delivery_date,:po_entry_date,
                 :data_json,:saved_by,:saved_at,:updated_at)
            ON CONFLICT(id) DO UPDATE SET
                wo=excluded.wo, customer=excluded.customer,
                division=excluded.division, priority=excluded.priority,
                coating=excluded.coating, electrode=excluded.electrode,
                order_qty=excluded.order_qty, order_value=excluded.order_value,
                po_advance=excluded.po_advance, trail=excluded.trail,
                design=excluded.design, store_blanks=excluded.store_blanks,
                shortage=excluded.shortage, bin_number=excluded.bin_number,
                rcvd_status=excluded.rcvd_status, rcvd_date=excluded.rcvd_date,
                rcvd_qty=excluded.rcvd_qty, del_start=excluded.del_start,
                del_weeks=excluded.del_weeks, delivery_date=excluded.delivery_date,
                po_entry_date=excluded.po_entry_date,
                data_json=excluded.data_json, saved_by=excluded.saved_by,
                updated_at=excluded.updated_at
        """, {
            "id":            record.get("id"),
            "wo":            record.get("id", ""),
            "customer":      customer_name,
            "division":      record.get("division", ""),
            "priority":      record.get("priority", ""),
            "coating":       record.get("coating", ""),
            "electrode":     record.get("electrode", ""),
            "order_qty":     record.get("orderQty"),
            "order_value":   record.get("orderValue"),
            "po_advance":    record.get("poAdvance", ""),
            "trail":         record.get("trail", ""),
            "design":        record.get("design", ""),
            "store_blanks":  record.get("storeAQTY"),
            "shortage":      record.get("shortage"),
            "bin_number":    record.get("binNumber", ""),
            "rcvd_status":   record.get("rcvdShortageStatus", ""),
            "rcvd_date":     record.get("rcvdShortageDate", ""),
            "rcvd_qty":      record.get("rcvdShortageQty"),
            "del_start":     record.get("woIssueDate", ""),
            "del_weeks":     record.get("weeks"),
            "delivery_date": record.get("deliveryDate", ""),
            "po_entry_date": record.get("poEntryDate", ""),
            "data_json":     json.dumps(record),
            "saved_by":      record.get("savedBy", ""),
            "saved_at":      record.get("savedAt", now),
            "updated_at":    now,
        })
        if customer_name:
            _upsert_customer(conn, customer_name)
        conn.commit()
        socketio.emit("data_changed", {"action": "saved", "id": record["id"]})
        if customer_name:
            socketio.emit("customers_changed", {})
        return jsonify({"ok": True, "id": record["id"]})
    finally:
        conn.close()


@app.route("/api/orders/<path:order_id>/attachment", methods=["POST"])
@login_required
def upload_attachment(order_id):
    _ensure_attachment_schema()
    _ensure_upload_folder()
    upload = request.files.get("file")
    if not upload or not upload.filename:
        return jsonify({"error": "No PDF file was provided"}), 400

    original_name = secure_filename(os.path.basename(upload.filename))
    if not original_name or not original_name.lower().endswith(".pdf"):
        return jsonify({"error": "Only PDF files are allowed"}), 400

    # Read actual bytes to enforce size limit (Content-Length header can be spoofed)
    pdf_data = upload.stream.read(MAX_UPLOAD_BYTES + 1)
    if len(pdf_data) > MAX_UPLOAD_BYTES:
        return jsonify({"error": "PDF must be 10 MB or smaller"}), 413

    conn = get_db()
    try:
        row = conn.execute(
            "SELECT attachment_path FROM work_orders WHERE id = ?",
            (order_id,)
        ).fetchone()
        if not row:
            return jsonify({"error": "Save the work order before attaching a PDF"}), 404

        stamp = datetime.datetime.utcnow().strftime("%Y%m%d%H%M%S")
        stored_name = f"{secure_filename(order_id) or 'order'}__{stamp}___{original_name}"
        with open(os.path.join(UPLOAD_FOLDER, stored_name), "wb") as _pdf_f:
            _pdf_f.write(pdf_data)
        _remove_attachment_file(row["attachment_path"])
        conn.execute(
            "UPDATE work_orders SET attachment_path = ?, updated_at = ? WHERE id = ?",
            (stored_name, datetime.datetime.utcnow().isoformat(timespec="seconds"), order_id)
        )
        conn.commit()
        socketio.emit("data_changed", {"action": "attachment_saved", "id": order_id})
        return jsonify({
            "ok": True,
            "attachmentPath": stored_name,
            "filename": _attachment_display_name(stored_name),
        })
    finally:
        conn.close()


@app.route("/api/orders/<path:order_id>/attachment", methods=["GET"])
@login_required
def get_attachment(order_id):
    _ensure_attachment_schema()
    conn = get_db()
    try:
        row = conn.execute(
            "SELECT attachment_path FROM work_orders WHERE id = ?",
            (order_id,)
        ).fetchone()
    finally:
        conn.close()

    if not row or not row["attachment_path"]:
        return jsonify({"error": "No PDF is attached to this work order"}), 404

    stored_name = os.path.basename(row["attachment_path"])
    file_path = os.path.join(UPLOAD_FOLDER, stored_name)
    if not os.path.exists(file_path):
        return jsonify({"error": "Attached PDF file was not found"}), 404

    return send_from_directory(
        UPLOAD_FOLDER,
        stored_name,
        mimetype="application/pdf",
        as_attachment=False,
    )


@app.route("/api/orders/<path:order_id>/attachment", methods=["DELETE"])
@login_required
def delete_attachment(order_id):
    _ensure_attachment_schema()
    conn = get_db()
    try:
        row = conn.execute(
            "SELECT attachment_path FROM work_orders WHERE id = ?",
            (order_id,)
        ).fetchone()
        if not row:
            return jsonify({"error": "Work order not found"}), 404
        if not row["attachment_path"]:
            return jsonify({"error": "No PDF is attached to this work order"}), 404

        stored_name = row["attachment_path"]
        conn.execute(
            "UPDATE work_orders SET attachment_path = NULL, updated_at = ? WHERE id = ?",
            (datetime.datetime.utcnow().isoformat(timespec="seconds"), order_id)
        )
        conn.commit()
        _remove_attachment_file(stored_name)
        socketio.emit("data_changed", {"action": "attachment_deleted", "id": order_id})
        return jsonify({"ok": True})
    finally:
        conn.close()


@app.route("/api/orders/<path:order_id>", methods=["DELETE"])
@login_required
def delete_order(order_id):
    user_key = session.get('user_key')
    conn = get_db()
    try:
        user = conn.execute(
            "SELECT is_admin FROM app_users WHERE user_key = ?", (user_key,)
        ).fetchone()
        if not user or not user['is_admin']:
            return jsonify({'error': 'Only administrators can delete orders'}), 403
        row = conn.execute(
            "SELECT attachment_path FROM work_orders WHERE id = ?",
            (order_id,)
        ).fetchone()
        conn.execute("DELETE FROM work_orders WHERE id = ?", (order_id,))
        conn.commit()
        if row:
            _remove_attachment_file(row["attachment_path"])
        socketio.emit("data_changed", {"action": "deleted", "id": order_id})
        return jsonify({"ok": True})
    finally:
        conn.close()

if __name__ == "__main__":
    port = int(os.getenv("PORT", "5000") or "5000")
    # Run without eventlet monkey-patching to avoid application/context issues.
    socketio.run(app, host="0.0.0.0", port=port, debug=False,
                 use_reloader=False, allow_unsafe_werkzeug=True)
